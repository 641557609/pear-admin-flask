import os
import openpyxl
from copy import copy
from collections import defaultdict


def split_excel_by_column(file_path, column_name):
    # 单次加载模板工作簿
    template_wb = openpyxl.load_workbook(file_path)
    sheets_info = {}
    person_data = defaultdict(lambda: defaultdict(set))

    # 第一阶段：收集数据和样式信息
    for sheet in template_wb:
        # 查找目标列
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            continue

        col_idx = None
        for cell in header_row:
            if cell.value == column_name:
                col_idx = cell.column
                break
        if not col_idx:
            continue

        sheets_info[sheet.title] = col_idx

        # 记录有效行号
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            person = str(row[col_idx - 1].value or "")
            if person:
                person_data[person][sheet.title].add(row_idx)

    if not person_data:
        template_wb.close()
        raise ValueError(f"文件中未找到列名 '{column_name}'")

    # 准备输出信息
    original_dir = os.path.dirname(file_path)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output = []

    # 第二阶段：生成结果文件
    for person, sheet_rows in person_data.items():
        safe_person = "".join(c if c.isalnum() else '_' for c in person)
        new_path = os.path.join(original_dir, f"{base_name}_{safe_person}.xlsx")

        # 创建新工作簿并复制样式模板
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 移除默认sheet

        for sheet_name in sheets_info:
            if sheet_name not in template_wb:
                continue

            # 复制原始工作表
            src_sheet = template_wb[sheet_name]
            dest_sheet = new_wb.create_sheet(sheet_name)

            # 复制列宽
            for col, dim in src_sheet.column_dimensions.items():
                dest_sheet.column_dimensions[col] = copy(dim)

            # 复制表头（带样式）
            for src_row in src_sheet.iter_rows(max_row=1):
                for cell in src_row:
                    dest_cell = dest_sheet.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value
                    )
                    if cell.has_style:
                        # 复制所有样式属性
                        dest_cell.font = copy(cell.font)
                        dest_cell.border = copy(cell.border)
                        dest_cell.fill = copy(cell.fill)
                        dest_cell.number_format = copy(cell.number_format)
                        dest_cell.protection = copy(cell.protection)
                        dest_cell.alignment = copy(cell.alignment)

            # 复制需要保留的行（带样式）
            dest_row_num = 2  # 从第二行开始写入数据
            for row_idx in sorted(sheet_rows.get(sheet_name, set())):
                src_row = src_sheet[row_idx]

                # 创建新行并复制数据
                for cell in src_row:
                    dest_cell = dest_sheet.cell(
                        row=dest_row_num,
                        column=cell.column,
                        value=cell.value
                    )
                    # 复制完整样式
                    if cell.has_style:
                        dest_cell.font = copy(cell.font)
                        dest_cell.border = copy(cell.border)
                        dest_cell.fill = copy(cell.fill)
                        dest_cell.number_format = cell.number_format
                        dest_cell.alignment = copy(cell.alignment)
                # 复制行高到新行位置
                dest_sheet.row_dimensions[dest_row_num].height = src_sheet.row_dimensions[row_idx].height
                dest_row_num += 1
            # 删除原始空行模板（原代码中多余的保留行）
            if dest_sheet.max_row > dest_row_num - 1:
                for row_idx in reversed(range(dest_row_num, dest_sheet.max_row + 1)):
                    dest_sheet.delete_rows(row_idx)

        new_wb.save(new_path)
        new_wb.close()
        output.append({"接收人": person, "文件名": new_path})

    template_wb.close()
    return output

# def split_excel_by_column(file_path, column_name):
#     # 单次加载模板工作簿
#     template_wb = openpyxl.load_workbook(file_path)
#     sheets_info = {}
#     person_data = defaultdict(lambda: defaultdict(set))
#
#     # 第一阶段：收集数据和样式信息
#     for sheet in template_wb:
#         # 查找目标列
#         header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
#         if not header_row:
#             continue
#
#         col_idx = None
#         for cell in header_row:
#             if cell.value == column_name:
#                 col_idx = cell.column
#                 break
#         if not col_idx:
#             continue
#
#         sheets_info[sheet.title] = col_idx
#
#         # 记录有效行号
#         for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
#             person = str(row[col_idx - 1].value or "")
#             if person:
#                 person_data[person][sheet.title].add(row_idx)
#
#     if not person_data:
#         template_wb.close()
#         raise ValueError(f"文件中未找到列名 '{column_name}'")
#
#     # 准备输出信息
#     original_dir = os.path.dirname(file_path)
#     base_name = os.path.splitext(os.path.basename(file_path))[0]
#     output = []
#
#     # 第二阶段：生成结果文件
#     for person, sheet_rows in person_data.items():
#         safe_person = "".join(c if c.isalnum() else '_' for c in person)
#         new_path = os.path.join(original_dir, f"{base_name}_{safe_person}.xlsx")
#
#         # 创建新工作簿并复制样式模板
#         new_wb = openpyxl.Workbook()
#         new_wb.remove(new_wb.active)  # 移除默认sheet
#
#         for sheet_name in sheets_info:
#             if sheet_name not in template_wb:
#                 continue
#
#             # 复制原始工作表
#             src_sheet = template_wb[sheet_name]
#             dest_sheet = new_wb.create_sheet(sheet_name)
#
#             # 复制列宽
#             for col, dim in src_sheet.column_dimensions.items():
#                 dest_sheet.column_dimensions[col] = copy(dim)
#
#             # 复制表头（带样式）
#             for src_row in src_sheet.iter_rows(max_row=1):
#                 for cell in src_row:
#                     dest_cell = dest_sheet.cell(
#                         row=cell.row,
#                         column=cell.column,
#                         value=cell.value
#                     )
#                     if cell.has_style:
#                         # 复制所有样式属性
#                         dest_cell.font = copy(cell.font)
#                         dest_cell.border = copy(cell.border)
#                         dest_cell.fill = copy(cell.fill)
#                         dest_cell.number_format = copy(cell.number_format)
#                         dest_cell.protection = copy(cell.protection)
#                         dest_cell.alignment = copy(cell.alignment)
#
#             # 复制需要保留的行（带样式）
#             dest_row_num = 2  # 从第二行开始写入数据
#             for row_idx in sorted(sheet_rows.get(sheet_name, set())):
#                 src_row = src_sheet[row_idx]
#
#                 # 创建新行并复制数据
#                 for cell in src_row:
#                     dest_cell = dest_sheet.cell(
#                         row=dest_row_num,
#                         column=cell.column,
#                         value=cell.value
#                     )
#                     # 复制完整样式
#                     if cell.has_style:
#                         dest_cell.font = copy(cell.font)
#                         dest_cell.border = copy(cell.border)
#                         dest_cell.fill = copy(cell.fill)
#                         dest_cell.number_format = cell.number_format
#                         dest_cell.alignment = copy(cell.alignment)
#                 # 复制行高到新行位置
#                 dest_sheet.row_dimensions[dest_row_num].height = src_sheet.row_dimensions[row_idx].height
#                 dest_row_num += 1
#             # 删除原始空行模板（原代码中多余的保留行）
#             if dest_sheet.max_row > dest_row_num - 1:
#                 for row_idx in reversed(range(dest_row_num, dest_sheet.max_row + 1)):
#                     dest_sheet.delete_rows(row_idx)
#
#         new_wb.save(new_path)
#         new_wb.close()
#         output.append({"接收人": person, "文件名": new_path})
#
#     template_wb.close()
#     return output
