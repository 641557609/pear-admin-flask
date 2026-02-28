# 用于生成Excel并存储的模块
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path
import logging
logger = logging.getLogger(__name__)



class ExcelGenerator:
    # 样式配置（可扩展）
    # 表头样式
    HEADER_STYLE = {
        'fill': PatternFill(start_color="C7E4B3", fill_type="solid"),
        'font': Font(bold=True, size=9, name='微软雅黑'),
        'alignment': Alignment(horizontal="center", vertical="center")
    }
    # 内容样式
    DATA_STYLE = {
        'font': Font(size=9, name='微软雅黑'),
        'alignment': Alignment(horizontal="center", vertical="justify")
    }

    def __init__(self, output_dir: str = "static/upload"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
            self,
            datas: Dict[str, List[List[Dict[str, list]]]],
            filename: str = None,
            max_col_width: int = 60,
            template_path: Optional[str] = None
    ) -> Path:
        """
        生成Excel文件核心方法

        参数:
        results: {
            "Sheet1": [
                {"headers": ["ID", "Name"]},
                {"data": [(1, "Alice"), (2, "Bob")]}
            ]
        }
        filename: 指定文件名（自动添加时间戳）
        max_col_width: 最大列宽限制
        template_path: 模板文件路径（可选），如果有模板则在模板基础上新建sheet，否则创建新文件
        """
        if template_path:
            template_path = Path(__file__).resolve().parent.parent.parent / template_path[1:]

        # 记录模板原有的sheet页名称
        original_sheets = []

        # 如果有模板文件路径，则基于模板创建
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            original_sheets = wb.sheetnames.copy() # 保存模板原有的sheet页名称
            logger.info(f"基于模板文件创建Excel: {template_path}")
        else:
            wb = Workbook()
            self._remove_default_sheet(wb)
            if template_path:
                logger.warning(f"模板文件不存在，将创建新的Excel文件: {template_path}")
        # 生成新的sheet页
        new_sheets = []
        for sheet_name, content in datas.items():
            if sheet_name in wb.sheetnames:
                # 如果工作表已存在，创建带时间戳的新工作表名
                timestamp = datetime.now().strftime('%H%M%S')
                new_sheet_name = f"{sheet_name}_{timestamp}"
                logger.info(f"工作表 '{sheet_name}' 已存在，将创建新工作表: {new_sheet_name}")
                ws = self._create_worksheet(wb, new_sheet_name)
                new_sheets.append(new_sheet_name)
            else:
                ws = self._create_worksheet(wb, sheet_name)
                new_sheets.append(sheet_name)

            self._process_sheet_content(ws, content, max_col_width)

            # 如果有模板，将模板原有的sheet页移动到所有新sheet页的后面
        if original_sheets:
            self._move_original_sheets_to_end(wb, original_sheets, new_sheets)

        return self._save_workbook(wb, filename)

    def _create_worksheet(self, wb: Workbook, title: str):
        """创建并激活工作表"""
        ws = wb.create_sheet(title=title)
        wb.active = ws
        return ws

    def _process_sheet_content(
            self,
            ws,
            content: List[List[Dict[str, list]]],
            max_col_width: int
    ):
        """处理单个工作表内容"""


        # 分离header和data
        for section_list in content:
            header_data = None
            all_data = []
            for section in section_list:
                if 'headers' in section:
                    header_data = section['headers']
                elif 'data' in section:
                    all_data.extend(section['data'])

            if header_data:
                self._write_headers(ws, header_data)
            if all_data:
                self._write_data(ws, all_data)
                self._auto_adjust_columns(ws, max_col_width)

    def _write_headers(self, ws, headers: list):
        """写入表头并应用样式"""
        ws.append(headers)
        for cell in ws[ws.max_row]:
            for style_type, style in self.HEADER_STYLE.items():
                setattr(cell, style_type, style)

    def _write_data(self, ws, data: list):
        """批量写入数据行"""
        for row in data:
            ws.append(list(row))
            self._apply_data_style(ws)

    def _apply_data_style(self, ws):
        """应用数据行样式"""
        for cell in ws[ws.max_row]:
            for style_type, style in self.DATA_STYLE.items():
                setattr(cell, style_type, style)

    def _auto_adjust_columns(self, ws, max_width: int):
        """智能列宽调整（带最大宽度限制）"""
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)

            for cell in col:
                cell_value = self._format_cell_value(cell.value)
                max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 7, max_width)
            ws.column_dimensions[column].width = adjusted_width

    def _format_cell_value(self, value):
        """格式化特殊数据类型"""
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        return value

    def _save_workbook(self, wb: Workbook, filename: str = None) -> Path:
        """保存文件并返回路径对象"""
        if not filename:
            filename = f"report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = self.output_dir / filename
        wb.save(file_path)
        return file_path

    @staticmethod
    def _remove_default_sheet(wb: Workbook):
        """移除默认的空白Sheet"""
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    def _move_original_sheets_to_end(self, wb: Workbook, original_sheets: list, new_sheets: list):
        """将模板原有的sheet页移动到所有新sheet页的后面"""
        # 获取当前所有sheet页的顺序
        current_order = wb.sheetnames

        # 创建新的sheet页顺序：新sheet页在前，模板原有sheet页在后
        new_order = []

        # 先添加所有新生成的sheet页（保持原有顺序）
        for sheet_name in current_order:
            if sheet_name in new_sheets:
                new_order.append(sheet_name)

        # 再添加模板原有的sheet页（保持原有顺序）
        for sheet_name in current_order:
            if sheet_name in original_sheets:
                new_order.append(sheet_name)

        # 重新排序sheet页
        wb._sheets.sort(key=lambda ws: new_order.index(ws.title))


